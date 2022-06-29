# 提示用户是否将BCA结果Excel表正确命名，并放于桌面
active = True
while active:
    print("输入quit，随时结束程序")
    confirm_data = input("请确认将BCA结果Excel表命名为：result.xlsx，并将其放于桌面，回答yes/no：")
    if confirm_data == 'quit':
        msg = "您选择了结束程序，现在关闭"
        print(msg)
        exit()
    elif confirm_data == 'yes':
        active = False

# 引入Excel读取数据库的库——openpyxl
import openpyxl

# 加载Excel
try:
    excel = openpyxl.load_workbook('/Users/zhaowang/Desktop/result.xlsx')
except FileNotFoundError:
    msg = "桌面没有命名为result的excel表，程序结束!"
    print(msg)
    exit()

# 当前激活的工作表
sheet = excel.active

data_tables = []
# 对BCA结果的单元格范围进行遍历
for spaces in sheet['C25':'N32']:
    for cell in spaces:
        if cell.value > 0:
            data_tables.append(cell.value)
        else:
            break

# 获得列表数据个数
nums = len(data_tables)

# 提示用户输入稀释倍数，默认为5倍
active = True
while active:
    folds = input("稀释倍数默认为10倍，回车即可，如更改请输入稀释倍数：")
    if folds == "":
        folds = 10
        active = False
    elif folds == 'quit':
        msg = "您选择了结束程序，现在关闭"
        print(msg)
        exit()
    else:
        try:
            folds = float(folds)
            if folds < 0:
                print("倍数不能为负值，请重新输入！")
            elif folds > 0:
                folds = float(folds)
                active = False
            elif folds == 0:
                folds = 1
                active = False
        except ValueError:
            print("无效输入，请输入数字倍数！")

# 提示用户输入上样量，默认为30µL
active = True
while active:
    s_volume = input("总上样量默认为30µL，回车即可，如更改请输入总上样量(µL)：")
    if s_volume == "":
        s_volume = 30
        active = False
    elif s_volume == 'quit':
        msg = "您选择了结束程序，现在关闭"
        print(msg)
        exit()
    else:
        try:
            s_volume = float(s_volume)
            if s_volume <= 0:
                print("总上样量不能为负值或零，请重新输入！")
            elif s_volume > 0:
                s_volume = float(s_volume)
                active = False
        except ValueError:
            print("无效输入，请重新输入总上样量！")

from openpyxl import Workbook

# 创建Excel表格
new_result = Workbook()
# 激活sheet页面
new_sheet = new_result.active
new_sheet_1 = new_result.create_sheet('result_sheet', 0)
# 写入表头
new_sheet_1['A1'] = '序号'
new_sheet_1['B1'] = 'OD值'
new_sheet_1['C1'] = '浓度(µg/µL)'
new_sheet_1['D1'] = '上样量(µL)'
new_sheet_1['E1'] = '裂解液补液量(µL)'
new_sheet_1['F1'] = '5X buffer量(µL)'
# 根据nums，将本次所有样品的序号写入Excel
num_lable = 1
while num_lable <= nums:
    num_labels = str(num_lable) + "号"
    num_lable = num_lable + 1
    A_numble = "A" + str(num_lable)
    new_sheet_1[A_numble] = num_labels

# 根据原始OD值换算各种数据
OD_values = []
CC_values = []
SC_values = []
PC_values = []
# 将原始OD值和换算后的浓度写入Excel相应位置
num_lable = 1
for num in data_tables:
    # 根据原始OD值、稀释倍数来换算浓度
    CC_value = ((num - 0.1613) / 1.005) * folds
    # 获取新的浓度列表
    CC_values.append(CC_value)
    num_lable = num_lable + 1
    B_numble = "B" + str(num_lable)
    C_numble = "C" + str(num_lable)
    new_sheet_1[B_numble] = num
    new_sheet_1[C_numble] = CC_value

# 将浓度值写入Excel相应位置
# 获取最小浓度值
CC_value_min = min(CC_values)
# 根据总上样量，写入相应的5X buffer的量
buffer_volume = s_volume / 4
# 获取上样量比值，将样品浓度除以最小浓度
num_lable = 1
for CC_value in CC_values:
    SC_value = s_volume * CC_value_min / CC_value
    PC_value = s_volume - SC_value
    SC_values.append(SC_value)
    num_lable = num_lable + 1
    D_numble = "D" + str(num_lable)
    E_numble = "E" + str(num_lable)
    F_numble = "F" + str(num_lable)
    new_sheet_1[D_numble] = round(SC_value, 1)
    new_sheet_1[E_numble] = round(PC_value, 1)
    new_sheet_1[F_numble] = buffer_volume

# 存储Excel表格于指定路径,并命名为new_excel.xlsx
new_result.save('/Users/zhaowang/Desktop/new_result.xlsx')
# 提示用户程序运行结束
msg = "BCA浓度结果将以new_result.xlsx形式输出到桌面，请查看！欢迎下次使用"
print(msg)
